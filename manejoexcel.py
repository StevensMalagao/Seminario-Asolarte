import openpyxl
import os

def agregar_datos_a_excel(nombre_archivo_excel, datos_a_agregar):

    try:
        if os.path.exists(nombre_archivo_excel):
            # Cargar el libro de trabajo existente
            libro = openpyxl.load_workbook(nombre_archivo_excel)
            hoja = libro.active
        else:
            # Crear un nuevo libro de trabajo y una hoja si el archivo no existe
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = "Datos Registrados" # Asigna un nombre a la hoja
            hoja.append(["Fecha y Hora", "Peso", "Material", "Reciclador", "Observaciones"]) # Encabezados de columna

        # Buscar la primera fila vacía
        # openpyxl.max_row obtiene el número total de filas, incluyendo las llenas.
        # Al usar append, se agregará automáticamente después de la última fila con datos.
        hoja.append(datos_a_agregar)

        # Guardar los cambios
        libro.save(nombre_archivo_excel)
        print(f"Datos agregados correctamente a '{nombre_archivo_excel}'.")

    except Exception as e:
        print(f"Ocurrió un error: {e}")
