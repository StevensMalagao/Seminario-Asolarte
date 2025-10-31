import serial
import serial.tools.list_ports
import time
import re
import openpyxl
import os
import customtkinter as ctk
from tkinter import messagebox
import threading
from PIL import Image, ImageDraw
import webbrowser

NOMBRE_ARCHIVO_EXCEL = "Registro.xlsx"
NOMBRE_BASE_DATOS_RECICLADORES = "Base_datos.xlsx"
VELOCIDAD_BAUDIOS = 115200
LOGO_PATH = "logoaso.png"
LOGO_SIZE = (220, 80)
REFRESH_ICON_PATH = "refresh_icon.png"

MATERIALES_POR_FAMILIA = {
    "METALES": {"ALUMINIO": 101, "CHATARRA": 102, "COBRE": 103, "BRONCE": 104, "ANTIMONIO": 105, "ACERO": 106, "OTROS METALES": 199},
    "PAPEL Y CARTÓN": {"ARCHIVO": 201, "CARTON": 202, "CUBETAS O PANELES": 203, "PERIODICO": 204, "PLEGADIZA": 205, "TETRA PACK": 206, "PLASTIFICADO": 207, "KRAF": 208, "OTROS PAPEL Y CARTÓN": 299},
    "PLÁSTICOS": {"ACRÍLICO": 301, "PASTA": 302, "PET": 303, "PVC": 304, "PLASTICO BLANCO": 305, "POLIETILENO": 306, "SOPLADO": 307, "POLIPROPILENO": 308, "OTROS PLÁSTICOS": 399},
    "VIDRIO": {"OTROS VIDRIOS": 499},
    "TEXTIL": {"OTROS TEXTILES": 599},
    "MADERA": {"OTROS MADERABLES": 699}
}

PESO_EMPAQUES = {
    "Ninguno": 0.0,
    "Estopa": 1.5,
    "Globo": 2.5,
    "Lona": 2.0
}

def gestionar_recicladores_excel(operacion, nombre_archivo, nombre_hoja="recicladores"):
    if not os.path.exists(nombre_archivo):
        try:
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = nombre_hoja
            hoja.append(["Nombre_reciclador", "ID_reciclador"])
            hoja.append(["Ejemplo Reciclador 1", "123456789"])
            hoja.append(["Ejemplo Reciclador 2", "987654321"])
            libro.save(nombre_archivo)
        except Exception as e:
            print(f"Error al crear el archivo Excel ('{nombre_archivo}'): {e}")
            return {"Error al crear archivo": "N/A"}

    try:
        libro = openpyxl.load_workbook(nombre_archivo)
        hoja = libro[nombre_hoja] if nombre_hoja in libro.sheetnames else libro.create_sheet(nombre_hoja)
        
        if hoja.max_row == 0:
            hoja.append(["Nombre_reciclador", "ID_reciclador"])
            hoja.append(["Ejemplo Reciclador 1", "123456789"])
            libro.save(nombre_archivo)

        if operacion == "leer":
            recicladores_db = {}
            for fila in range(2, hoja.max_row + 1):
                nombre_cell = hoja.cell(row=fila, column=1).value
                id_cell = hoja.cell(row=fila, column=2).value
                
                if nombre_cell:
                    nombre = str(nombre_cell).strip()
                    id_reciclador = str(id_cell).strip() if id_cell else "N/A"
                    recicladores_db[nombre] = id_reciclador
            
            return recicladores_db if recicladores_db else {"Lista vacía en Excel": "N/A"}

    except Exception as e:
        print(f"Error al operar con Excel ('{nombre_archivo}'): {e}")
        return {"Error al leer archivo": "N/A"}

def agregar_datos_a_excel(nombre_archivo_excel, datos_registro):
    fecha = datos_registro["fecha"]
    reciclador = datos_registro["reciclador"]
    reciclador_id = datos_registro["reciclador_id"]
    material = datos_registro["material"]
    peso_neto = float(datos_registro["peso_neto"])

    NOMBRE_HOJA = "Datos Agrupados por Dia"
    
    COL_FECHA_NOMBRE = "Fecha"
    COL_RECICLADOR_NOMBRE = "Reciclador"
    COL_ID_NOMBRE = "ID"
    COL_TOTAL_NOMBRE = "Total"
    BASE_HEADERS = [COL_FECHA_NOMBRE, COL_RECICLADOR_NOMBRE, COL_ID_NOMBRE, COL_TOTAL_NOMBRE]

    try:
        if os.path.exists(nombre_archivo_excel):
            libro = openpyxl.load_workbook(nombre_archivo_excel)
            hoja = libro[NOMBRE_HOJA] if NOMBRE_HOJA in libro.sheetnames else libro.create_sheet(NOMBRE_HOJA)
        else:
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = NOMBRE_HOJA

        if hoja.cell(row=1, column=1).value is None:
            print("Celda A1 vacía detectada. Escribiendo encabezados...")
            for idx, header_name in enumerate(BASE_HEADERS):
                hoja.cell(row=1, column=idx + 1).value = header_name

        headers = [cell.value for cell in hoja[1]]
        
        def get_or_create_col_idx(col_name, before_col_name=COL_TOTAL_NOMBRE):
            nonlocal headers
            try:
                return headers.index(col_name) + 1
            except ValueError:
                try:
                    total_col_idx_base_0 = headers.index(before_col_name)
                except ValueError:
                    total_col_idx_base_0 = len(headers)
                    hoja.cell(row=1, column=total_col_idx_base_0 + 1).value = before_col_name
                    headers.append(before_col_name)

                new_col_idx_base_0 = total_col_idx_base_0
                new_col_idx_base_1 = new_col_idx_base_0 + 1
                
                hoja.insert_cols(new_col_idx_base_1)
                hoja.cell(row=1, column=new_col_idx_base_1).value = col_name
                
                headers.insert(new_col_idx_base_0, col_name)
                
                return new_col_idx_base_1

        col_fecha_idx = get_or_create_col_idx(COL_FECHA_NOMBRE)
        col_reciclador_idx = get_or_create_col_idx(COL_RECICLADOR_NOMBRE)
        col_id_idx = get_or_create_col_idx(COL_ID_NOMBRE)
        col_material_idx = get_or_create_col_idx(material)
        col_total_idx = get_or_create_col_idx(COL_TOTAL_NOMBRE)

        target_row_num = -1
        for row_idx in range(2, hoja.max_row + 1):
            if hoja.cell(row=row_idx, column=col_fecha_idx).value == fecha and \
               hoja.cell(row=row_idx, column=col_reciclador_idx).value == reciclador:
                target_row_num = row_idx
                break
        
        if target_row_num == -1:
            target_row_num = hoja.max_row + 1
            hoja.cell(row=target_row_num, column=col_fecha_idx).value = fecha
            hoja.cell(row=target_row_num, column=col_reciclador_idx).value = reciclador
            hoja.cell(row=target_row_num, column=col_id_idx).value = reciclador_id
            
            for idx, header_name in enumerate(headers):
                if header_name and header_name not in [COL_FECHA_NOMBRE, COL_RECICLADOR_NOMBRE, COL_ID_NOMBRE]:
                    hoja.cell(row=target_row_num, column=idx + 1).value = 0.0

        current_material_val_cell = hoja.cell(row=target_row_num, column=col_material_idx)
        try:
            current_material_val = float(current_material_val_cell.value or 0.0)
        except ValueError:
            print(f"Advertencia: Celda de material contenía texto '{current_material_val_cell.value}'. Reiniciando a 0.")
            current_material_val = 0.0
        current_material_val_cell.value = current_material_val + peso_neto
        
        current_total_val_cell = hoja.cell(row=target_row_num, column=col_total_idx)
        try:
            current_total_val = float(current_total_val_cell.value or 0.0)
        except ValueError:
            print(f"Advertencia: Celda total contenía texto '{current_total_val_cell.value}'. Reiniciando a 0.")
            current_total_val = 0.0
        current_total_val_cell.value = current_total_val + peso_neto

        libro.save(nombre_archivo_excel)
        return True
    except Exception as e:
        print(f"Ocurrió un error al escribir en Excel (formato agrupado): {e}")
        messagebox.showerror("Error al Guardar en Excel", f"No se pudo guardar el registro agrupado en '{NOMBRE_HOJA}'.\n\nError: {e}\n\nAsegúrese de que el archivo no esté abierto por otro programa.")
        return False

def manejar_error_serial(gui_app, puerto):
    if gui_app.running:
        messagebox.showerror("Error de Conexión", f"Se perdió la conexión con {puerto}.\nVerifique el dispositivo.\nLa aplicación se cerrará.")
        if gui_app.winfo_exists():
            gui_app.destroy()

def leer_serial(puerto, baudios, gui_app):
    ser = None
    try:
        ser = serial.Serial(puerto, baudios, timeout=1)
        print(f"Abriendo puerto serial {puerto} a {baudios} baudios...")
        time.sleep(2)
        while gui_app.running:
            if ser.in_waiting > 0:
                linea = ser.readline().decode('utf-8').strip()
                if linea: gui_app.procesar_linea_serial(linea)
            time.sleep(0.1)
    except serial.SerialException:
        print(f"Error de conexión en puerto {puerto}.")
        if gui_app.running:
            gui_app.after(0, manejar_error_serial, gui_app, puerto)
    except Exception as e:
        print(f"Ocurrió un error inesperado en el hilo serial: {e}")
    finally:
        if ser and ser.is_open: ser.close()
        print("Puerto serial cerrado.")

class AplicacionReciclaje(ctk.CTk):
    def __init__(self, puerto_com_seleccionado):
        super().__init__()
        
        self.peso_lock = threading.Lock()
        self.puerto_serial = puerto_com_seleccionado
        self.peso_bruto_actual = 0.0
        self.running = True
        
        self.recicladores_db = gestionar_recicladores_excel("leer", NOMBRE_BASE_DATOS_RECICLADORES)
        lista_nombres_recicladores = sorted(list(self.recicladores_db.keys()))

        self.title(f"Sistema de Registro Asorate - Conectado a {self.puerto_serial}")
        self.geometry("800x650") 
        self.resizable(False, False)
        self.grid_columnconfigure(0, weight=1)
        
        self.fecha_var = ctk.StringVar(value="Esperando...")
        self.peso_bruto_var = ctk.StringVar(value="0.00 kg")
        self.tara_var = ctk.StringVar(value="0.00 kg")
        self.peso_neto_var = ctk.StringVar(value="0.00 kg")
        self.empaque_seleccionado = ctk.StringVar(value="Ninguno")
        self.familia_seleccionada = ctk.StringVar(value="Seleccione una familia")
        self.material_seleccionado = ctk.StringVar(value="Seleccione un material")
        self.reciclador_seleccionado = ctk.StringVar(value="Seleccione un reciclador")
        self.novedad_var = ctk.StringVar()

        self.refresh_icon = None
        if os.path.exists(REFRESH_ICON_PATH):
            try:
                self.refresh_icon = ctk.CTkImage(Image.open(REFRESH_ICON_PATH), size=(20, 20))
            except Exception as e:
                print(f"Error al cargar el icono de refrescar: {e}")
        else:
            print(f"Advertencia (AplicacionReciclaje): No se encontró el icono '{REFRESH_ICON_PATH}'. Se usará texto.")

        self.create_widgets(lista_nombres_recicladores)
        self.iniciar_lectura_serial()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_widgets(self, lista_recicladores):
        if os.path.exists(LOGO_PATH):
            try:
                ctk.CTkLabel(self, image=ctk.CTkImage(Image.open(LOGO_PATH), size=LOGO_SIZE), text="").grid(row=0, column=0, padx=20, pady=(10, 5), sticky="n")
            except Exception as e: print(f"Error al cargar el logo: {e}")
        else: print(f"Advertencia: Archivo de logo '{LOGO_PATH}' no encontrado.")
        
        data_frame = ctk.CTkFrame(self)
        data_frame.grid(row=1, column=0, padx=20, pady=15, sticky="ew")
        data_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(data_frame, text="Fecha:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.fecha_var).grid(row=0, column=1, padx=10, pady=5, sticky="w")
        
        ctk.CTkLabel(data_frame, text="Peso Bruto:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.peso_bruto_var).grid(row=2, column=1, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, text="Tara (Empaque):", font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.tara_var).grid(row=3, column=1, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, text="Peso Neto:", font=ctk.CTkFont(weight="bold")).grid(row=4, column=0, padx=10, pady=10, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.peso_neto_var).grid(row=4, column=1, padx=10, pady=10, sticky="w")

        input_frame = ctk.CTkFrame(self)
        input_frame.grid(row=2, column=0, padx=20, pady=10, sticky="nsew")
        input_frame.grid_columnconfigure((1, 3), weight=1)

        ctk.CTkLabel(input_frame, text="Familia de Material:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        ctk.CTkComboBox(input_frame, variable=self.familia_seleccionada, values=list(MATERIALES_POR_FAMILIA.keys()), state="readonly", command=self.actualizar_materiales_por_familia).grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        ctk.CTkLabel(input_frame, text="Material:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.material_menu = ctk.CTkComboBox(input_frame, variable=self.material_seleccionado, values=[], state="disabled")
        self.material_menu.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
        ctk.CTkLabel(input_frame, text="Tipo de Empaque:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        ctk.CTkComboBox(input_frame, variable=self.empaque_seleccionado, values=list(PESO_EMPAQUES.keys()), state="readonly", command=self.actualizar_pesos).grid(row=2, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text="Reciclador:").grid(row=0, column=2, padx=(20, 10), pady=10, sticky="w")
        reciclador_frame = ctk.CTkFrame(input_frame, fg_color="transparent")
        reciclador_frame.grid(row=0, column=3, padx=10, pady=10, sticky="ew")
        reciclador_frame.grid_columnconfigure(0, weight=1)
        
        self.reciclador_menu = ctk.CTkComboBox(reciclador_frame, variable=self.reciclador_seleccionado, values=lista_recicladores, state="readonly")
        self.reciclador_menu.grid(row=0, column=0, sticky="ew")
        
        refresh_text = "" if self.refresh_icon else "Refrescar"
        refresh_width = 35 if self.refresh_icon else 80
        
        ctk.CTkButton(
            reciclador_frame, 
            text=refresh_text, 
            image=self.refresh_icon, 
            width=refresh_width,
            command=self.actualizar_dropdown_recicladores,
            fg_color="gray", 
            hover_color="darkgray"
        ).grid(row=0, column=1, padx=(10,0))
        
        ctk.CTkLabel(input_frame, text="Novedad:").grid(row=1, column=2, padx=(20, 10), pady=10, sticky="w")
        ctk.CTkEntry(input_frame, textvariable=self.novedad_var).grid(row=1, column=3, rowspan=2, padx=10, pady=10, sticky="nsew")

        ctk.CTkButton(self, text="Registrar Reciclaje", command=self.guardar_datos_gui, font=ctk.CTkFont(weight="bold"), height=40).grid(row=3, column=0, padx=20, pady=(10, 5), sticky="ew")
        ctk.CTkButton(self, text="Abrir Archivo de Registros", command=self.abrir_excel_registros, height=40, fg_color="#555555", hover_color="#444444").grid(row=4, column=0, padx=20, pady=(5, 5), sticky="ew")
        ctk.CTkButton(self, text="Modificar DB Recicladores", command=self.abrir_excel_base_de_datos, height=40, fg_color="#555555", hover_color="#444444").grid(row=5, column=0, padx=20, pady=(5, 20), sticky="ew")

    def abrir_excel_base_de_datos(self):
        gestionar_recicladores_excel("leer", NOMBRE_BASE_DATOS_RECICLADORES)
        try:
            webbrowser.open(os.path.abspath(NOMBRE_BASE_DATOS_RECICLADORES))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo '{NOMBRE_BASE_DATOS_RECICLADORES}':\n{e}")

    def actualizar_dropdown_recicladores(self):
        print("Actualizando lista de recicladores...")
        self.recicladores_db = gestionar_recicladores_excel("leer", NOMBRE_BASE_DATOS_RECICLADORES)
        nuevos_recicladores = sorted(list(self.recicladores_db.keys()))
        
        self.reciclador_menu.configure(values=nuevos_recicladores)
        self.reciclador_seleccionado.set("Seleccione un reciclador")
        print("Lista de recicladores actualizada.")
        messagebox.showinfo("Actualizado", "La lista de recicladores ha sido actualizada desde el archivo Excel.")

    def actualizar_materiales_por_familia(self, familia):
        materiales = list(MATERIALES_POR_FAMILIA.get(familia, {}).keys())
        self.material_menu.configure(values=materiales, state="readonly" if materiales else "disabled")
        self.material_seleccionado.set("Seleccione un material" if materiales else "No hay materiales")

    def iniciar_lectura_serial(self):
        self.serial_thread = threading.Thread(target=leer_serial, args=(self.puerto_serial, VELOCIDAD_BAUDIOS, self), daemon=True)
        self.serial_thread.start()
    
    def procesar_linea_serial(self, linea):
        full_match = re.match(r'Fecha: ([\d/]+)(?:, Peso: ([\d.]+))?', linea)
        
        if full_match:
            fecha, peso_str = full_match.groups()
            peso_actual = float(peso_str) if peso_str else self.peso_bruto_actual
            self.update_serial_data(fecha, peso_actual)
        elif re.match(r'^([\d.]+)$', linea):
            self.update_serial_data(self.fecha_var.get(), float(linea))

    def update_serial_data(self, fecha, peso_bruto):
        with self.peso_lock:
            self.peso_bruto_actual = float(peso_bruto)
        
        self.after(0, self._update_gui_labels, fecha)

    def _update_gui_labels(self, fecha):
        if fecha != "Esperando...": self.fecha_var.set(fecha)
        self.actualizar_pesos()
        
    def actualizar_pesos(self, _=None):
        tara = PESO_EMPAQUES.get(self.empaque_seleccionado.get(), 0.0)
        
        with self.peso_lock:
            peso_bruto_local = self.peso_bruto_actual
            
        self.peso_bruto_var.set(f"{peso_bruto_local:.2f} kg")
        self.tara_var.set(f"{tara:.2f} kg")
        self.peso_neto_var.set(f"{peso_bruto_local - tara:.2f} kg")

    def guardar_datos_gui(self):
        peso_neto_str = self.peso_neto_var.get().replace(" kg", "")
        nombre_reciclador = self.reciclador_seleccionado.get()
        
        with self.peso_lock:
            peso_bruto_local = self.peso_bruto_actual
        
        reciclador_id = self.recicladores_db.get(nombre_reciclador, "N/A")
            
        datos = {
            "fecha": self.fecha_var.get(),
            "peso_bruto": peso_bruto_local,
            "empaque": self.empaque_seleccionado.get(),
            "tara": PESO_EMPAQUES.get(self.empaque_seleccionado.get(), 0.0),
            "peso_neto": float(peso_neto_str),
            "familia": self.familia_seleccionada.get(),
            "material": self.material_seleccionado.get(),
            "reciclador": nombre_reciclador,
            "reciclador_id": reciclador_id,
            "novedad": self.novedad_var.get() or "N/A"
        }
        
        if any(v in ("Esperando...", "N/A") for v in (datos["fecha"],)) or any("Seleccione" in v or "Error" in v or "vacía" in v for v in (datos["familia"], datos["material"], datos["reciclador"])):
            messagebox.showwarning("Datos Incompletos", "Por favor, complete todos los campos de selección y espere los datos de la báscula."); return
        
        if datos["peso_neto"] <= 0 and not messagebox.askyesno("Peso Negativo o Cero", f"El peso neto es {datos['peso_neto']:.2f} kg. ¿Desea registrarlo de todas formas?"): return

        if agregar_datos_a_excel(NOMBRE_ARCHIVO_EXCEL, datos):
            messagebox.showinfo("Registro Exitoso", "Datos guardados en Excel correctamente.")
            self.reset_fields()
        else: 
            print("Error al guardar. El messagebox de error debe haberse mostrado.")

    def reset_fields(self):
        self.novedad_var.set("")
        self.familia_seleccionada.set("Seleccione una familia")
        self.material_menu.configure(values=[], state="disabled")
        self.material_seleccionado.set("Seleccione un material")
        self.reciclador_seleccionado.set("Seleccione un reciclador")
        self.empaque_seleccionado.set("Ninguno")
        self.fecha_var.set("Esperando...")
        
        with self.peso_lock:
            self.peso_bruto_actual = 0.0
            
        self.actualizar_pesos()

    def abrir_excel_registros(self):
        if not os.path.exists(NOMBRE_ARCHIVO_EXCEL): 
            messagebox.showwarning("Archivo no encontrado", f"'{NOMBRE_ARCHIVO_EXCEL}' aún no existe. Se creará con el primer registro."); return
        try: 
            webbrowser.open(os.path.abspath(NOMBRE_ARCHIVO_EXCEL))
        except Exception as e: 
            messagebox.showerror("Error", f"Ocurrió un error al abrir el archivo: {e}")

    def on_closing(self):
        if messagebox.askokcancel("Cerrar", "¿Deseas cerrar la aplicación?"):
            self.running = False
            if hasattr(self, 'serial_thread') and self.serial_thread.is_alive(): 
                self.serial_thread.join(1)
            self.destroy()

class VentanaSeleccionCOM(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Asorate - Conectar Dispositivo")
        self.geometry("400x180")
        self.resizable(False, False)
        self.grid_columnconfigure(0, weight=1)

        self.refresh_icon = None
        if os.path.exists(REFRESH_ICON_PATH):
            try:
                self.refresh_icon = ctk.CTkImage(Image.open(REFRESH_ICON_PATH), size=(20, 20))
            except Exception as e:
                print(f"Error al cargar el icono de refrescar: {e}")
        else:
            print(f"Advertencia (VentanaSeleccionCOM): No se encontró el icono '{REFRESH_ICON_PATH}'. Se usará texto.")

        ctk.CTkLabel(self, text="Seleccione el puerto del dispositivo:", font=ctk.CTkFont(size=14)).pack(padx=20, pady=(20, 5))
        
        com_frame = ctk.CTkFrame(self, fg_color="transparent")
        com_frame.pack(padx=20, pady=5, fill="x")
        com_frame.grid_columnconfigure(0, weight=1)

        self.puerto_seleccionado = ctk.StringVar()
        self.combobox_puertos = ctk.CTkComboBox(com_frame, variable=self.puerto_seleccionado, state="readonly")
        self.combobox_puertos.grid(row=0, column=0, sticky="ew")

        refresh_text = "" if self.refresh_icon else "Refrescar"
        refresh_width = 35 if self.refresh_icon else 80
        
        ctk.CTkButton(
            com_frame, 
            text=refresh_text, 
            image=self.refresh_icon, 
            command=self.refrescar_puertos, 
            width=refresh_width,
            fg_color="gray", 
            hover_color="darkgray",
            height=28
        ).grid(row=0, column=1, padx=(10,0))
        
        self.refrescar_puertos()

        ctk.CTkButton(self, text="Conectar", command=self.iniciar_app_principal, height=35).pack(padx=20, pady=(10,20), fill="x", expand=True)

    def refrescar_puertos(self):
        puertos = [port.device for port in serial.tools.list_ports.comports()]
        if puertos:
            self.combobox_puertos.configure(values=puertos)
            self.puerto_seleccionado.set(puertos[0])
        else:
            self.combobox_puertos.configure(values=["No se encontraron puertos"])
            self.puerto_seleccionado.set("No se encontraron puertos")
        print("Puertos actualizados:", puertos if puertos else "Ninguno")

    def iniciar_app_principal(self):
        puerto_elegido = self.puerto_seleccionado.get()
        if "No se encontraron" in puerto_elegido:
            messagebox.showerror("Error", "No se encontró ningún puerto COM. Asegúrese de que el dispositivo esté conectado y refresque la lista.")
            return
        self.destroy()
        app = AplicacionReciclaje(puerto_com_seleccionado=puerto_elegido)
        app.mainloop()

if __name__ == "__main__":
    ctk.set_appearance_mode("Light")
    ctk.set_default_color_theme("dark-blue")
    
    if not os.path.exists(LOGO_PATH):
        try:
            img = Image.new('RGB', LOGO_SIZE, color = 'gray')
            d = ImageDraw.Draw(img)
            d.text((10,20), "ASORATE", fill=(0,0,0))
            img.save(LOGO_PATH)
            print(f"Logo de ejemplo '{LOGO_PATH}' creado.")
        except Exception as e: print(f"No se pudo crear el logo de ejemplo: {e}.")

    ventana_inicial = VentanaSeleccionCOM()
    ventana_inicial.mainloop()
